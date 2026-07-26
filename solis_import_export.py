import csv
import datetime as dt
import json
import re
from pathlib import Path

from openpyxl import load_workbook


IMPORT_DIR = Path("solis_imports")
OUTPUT_FILE = Path("solis_generation.json")
SUPPORTED_SUFFIXES = {".csv", ".xlsx"}


FIELD_PATTERNS = {
    "name": [
        r"system.*name",
        r"station.*name",
        r"plant.*name",
        r"site.*name",
        r"name",
    ],
    "status": [r"status", r"state"],
    "capacity_kw": [r"capacity", r"power.*kw", r"installed"],
    "today_generation_kwh": [
        r"today.*generation",
        r"daily.*generation",
        r"today.*yield",
        r"daily.*yield",
        r"eday",
    ],
    "weekly_generation_kwh": [
        r"week.*generation",
        r"weekly.*generation",
        r"week.*yield",
        r"weekly.*yield",
    ],
    "total_generation_kwh": [
        r"total.*generation",
        r"total.*yield",
        r"lifetime",
        r"etotal",
    ],
    "system_id": [r"source.*id", r"station.*id", r"plant.*id", r"system.*id", r"id"],
}


def normalize(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def number(value):
    if value is None or value == "":
        return None
    text = str(value).strip().replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def latest_import_file():
    IMPORT_DIR.mkdir(exist_ok=True)
    files = [
        path
        for path in IMPORT_DIR.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES
    ]
    if not files:
        raise RuntimeError(
            f"No Solis export found in {IMPORT_DIR}. Put a .csv or .xlsx file there first."
        )
    return max(files, key=lambda path: path.stat().st_mtime)


def rows_from_csv(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def rows_from_xlsx(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(values[:20]):
        labels = [normalize(cell) for cell in row if cell is not None]
        if any("name" in label for label in labels) and any(
            "generation" in label or "yield" in label or "capacity" in label
            for label in labels
        ):
            header_index = index
            break
    if header_index is None:
        header_index = 0

    headers = [str(cell or "").strip() for cell in values[header_index]]
    rows = []
    for row in values[header_index + 1 :]:
        item = {
            headers[column]: row[column] if column < len(row) else None
            for column in range(len(headers))
            if headers[column]
        }
        if any(value not in (None, "") for value in item.values()):
            rows.append(item)
    return rows


def map_headers(headers):
    mapping = {}
    normalized = {header: normalize(header) for header in headers}
    for field, patterns in FIELD_PATTERNS.items():
        for header, label in normalized.items():
            if any(re.search(pattern, label) for pattern in patterns):
                mapping[field] = header
                break
    return mapping


def convert_rows(source_rows):
    if not source_rows:
        return []
    headers = list(source_rows[0].keys())
    mapping = map_headers(headers)
    systems = []
    for row in source_rows:
        name = row.get(mapping.get("name", ""), "")
        if not name:
            continue
        systems.append(
            {
                "name": str(name).strip(),
                "status": str(row.get(mapping.get("status", ""), "") or "").strip(),
                "capacity_kw": number(row.get(mapping.get("capacity_kw", ""))),
                "today_generation_kwh": number(
                    row.get(mapping.get("today_generation_kwh", ""))
                ),
                "weekly_generation_kwh": number(
                    row.get(mapping.get("weekly_generation_kwh", ""))
                ),
                "total_generation_kwh": number(
                    row.get(mapping.get("total_generation_kwh", ""))
                ),
                "system_id": str(row.get(mapping.get("system_id", ""), "") or "").strip(),
            }
        )
    return systems


def main():
    source = latest_import_file()
    if source.suffix.lower() == ".csv":
        source_rows = rows_from_csv(source)
    else:
        source_rows = rows_from_xlsx(source)

    systems = convert_rows(source_rows)
    if not systems:
        raise RuntimeError(f"No Solis systems could be read from {source}")

    payload = {
        "source": str(source),
        "generated_at": dt.datetime.now().replace(microsecond=0).isoformat(),
        "systems": systems,
    }
    OUTPUT_FILE.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Saved {len(systems)} Solis systems to {OUTPUT_FILE}")
    print(f"Source file: {source}")


if __name__ == "__main__":
    main()
