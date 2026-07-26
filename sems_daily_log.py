import datetime as dt
import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

import sems_api


OUTPUT_FILE = Path("sems_daily_generation.xlsx")
HEADERS = [
    "Date",
    "Station Name",
    "Daily Generation (kWh)",
    "Total Generation (kWh)",
]


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def append_row(row: dict[str, object]) -> None:
    if OUTPUT_FILE.exists():
        workbook = load_workbook(OUTPUT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Daily Generation"
        sheet.append(HEADERS)

    sheet.append([row.get(header) for header in HEADERS])
    workbook.save(OUTPUT_FILE)


def main() -> None:
    username = require_env("SEMS_USERNAME")
    password = require_env("SEMS_PASSWORD")
    plant_id = require_env("SEMS_PLANT_ID")

    client = sems_api.SemsAPI(username=username, password=password)

    yesterday = dt.date.today() - dt.timedelta(days=1)
    plant_data = client.get_plant_historical_data(
        plant_id,
        date=yesterday.strftime("%Y-%m-%d"),
    )

    row = {
        "Date": yesterday.isoformat(),
        "Station Name": plant_data.get("plant_name"),
        "Daily Generation (kWh)": plant_data.get("generation_today"),
        "Total Generation (kWh)": plant_data.get("generation_total"),
    }

    append_row(row)
    print(f"Added row to {OUTPUT_FILE}: {row}")


if __name__ == "__main__":
    main()
