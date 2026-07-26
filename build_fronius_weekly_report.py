import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_FILE = Path("fronius_weekly_generation.json")
OUTPUT_DIR = Path("outputs/fronius_weekly_generation")
OUTPUT_FILE = OUTPUT_DIR / "fronius_weekly_generation_report.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="EEF5F1")
NOTE_FILL = PatternFill("solid", fgColor="F7F9FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)


def set_widths(sheet, widths):
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def style_table(sheet, min_row, max_row, min_col, max_col):
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def main() -> None:
    payload = json.loads(SOURCE_FILE.read_text(encoding="utf-8"))
    systems = payload["systems"]
    dates = [item["date"] for item in systems[0]["daily"]] if systems else []

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Weekly Report"
    data = workbook.create_sheet("Daily Data")

    summary.sheet_view.showGridLines = False
    data.sheet_view.showGridLines = False

    summary["A1"] = "Fronius Weekly Generation Report"
    summary["A1"].font = Font(bold=True, size=18, color="17324D")
    summary["A2"] = f"Period: {payload['start_date']} to {payload['end_date']}"
    summary["A2"].font = Font(color="576575")

    summary_headers = [
        "System Name",
        "Status",
        "Weekly Generation (kWh)",
        "Average Daily (kWh)",
        "Best Day (kWh)",
    ]
    summary.append([])
    summary.append(summary_headers)
    for cell in summary[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for system in sorted(
        systems, key=lambda item: item["weekly_generation_kwh"], reverse=True
    ):
        daily_values = [item["generation_kwh"] for item in system["daily"]]
        weekly = system["weekly_generation_kwh"]
        summary.append(
            [
                system["name"],
                system["status"],
                weekly,
                weekly / len(daily_values) if daily_values else 0,
                max(daily_values) if daily_values else 0,
            ]
        )

    last_summary_row = 4 + len(systems)
    style_table(summary, 4, last_summary_row, 1, 5)
    for row in summary.iter_rows(min_row=5, max_row=last_summary_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"

    summary["G4"] = "Total Weekly kWh"
    summary["H4"] = f"=SUM(C5:C{last_summary_row})"
    summary["G5"] = "Systems"
    summary["H5"] = len(systems)
    summary["G6"] = "Producing Systems"
    summary["H6"] = f"=COUNTIF(C5:C{last_summary_row},\">0\")"
    for row in summary.iter_rows(min_row=4, max_row=6, min_col=7, max_col=8):
        for cell in row:
            cell.fill = SUBTLE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    summary["G4"].font = summary["G5"].font = summary["G6"].font = Font(
        bold=True, color="17324D"
    )
    summary["H4"].number_format = "#,##0.00"

    chart = BarChart()
    chart.title = "Weekly Generation by Fronius System"
    chart.y_axis.title = "kWh"
    chart.x_axis.title = "System"
    chart.height = 9
    chart.width = 18
    values = Reference(summary, min_col=3, min_row=4, max_row=last_summary_row)
    categories = Reference(summary, min_col=1, min_row=5, max_row=last_summary_row)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    summary.add_chart(chart, "G8")

    data_headers = ["Date"] + [system["name"] for system in systems]
    data.append(data_headers)
    for cell in data[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for index, date_value in enumerate(dates):
        data.append(
            [date_value] + [system["daily"][index]["generation_kwh"] for system in systems]
        )

    last_data_row = 1 + len(dates)
    last_data_col = 1 + len(systems)
    style_table(data, 1, last_data_row, 1, last_data_col)
    for row in data.iter_rows(min_row=2, max_row=last_data_row, min_col=2, max_col=last_data_col):
        for cell in row:
            cell.number_format = "#,##0.00"

    line_chart = LineChart()
    line_chart.title = "Daily Generation Trend"
    line_chart.y_axis.title = "kWh"
    line_chart.x_axis.title = "Date"
    line_chart.height = 9
    line_chart.width = 22
    line_values = Reference(data, min_col=2, max_col=last_data_col, min_row=1, max_row=last_data_row)
    line_categories = Reference(data, min_col=1, min_row=2, max_row=last_data_row)
    line_chart.add_data(line_values, titles_from_data=True)
    line_chart.set_categories(line_categories)
    data.add_chart(line_chart, "A11")

    summary["A15"] = "Note"
    summary["A15"].font = Font(bold=True, color="17324D")
    summary["A16"] = (
        "This report uses Fronius Solar.web production chart data fetched through the logged-in browser session."
    )
    summary["A16"].fill = NOTE_FILL
    summary["A16"].alignment = Alignment(wrap_text=True)

    set_widths(
        summary,
        {"A": 28, "B": 16, "C": 22, "D": 20, "E": 16, "G": 20, "H": 16},
    )
    for col in range(1, last_data_col + 1):
        data.column_dimensions[get_column_letter(col)].width = 18 if col > 1 else 14

    summary.freeze_panes = "A5"
    data.freeze_panes = "B2"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"Saved report to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
