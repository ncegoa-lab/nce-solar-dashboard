import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


GOODWE_FILE = Path("sems_station_data.json")
GOODWE_WEEKLY_FILE = Path("sems_weekly_generation.json")
FRONIUS_WEEKLY_FILE = Path("fronius_weekly_generation.json")
FRONIUS_CURRENT_FILE = Path("fronius_current_generation.json")
FIMER_FILE = Path("fimer_generation.json")
SOLIS_FILE = Path("solis_generation.json")
SOLAX_FILE = Path("solax_generation.json")
OUTPUT_DIR = Path("outputs/numbers_compatible")
OUTPUT_FILE = OUTPUT_DIR / "numbers_compatible_generation_report.xlsx"
CSV_FILE = OUTPUT_DIR / "numbers_compatible_generation_report.csv"


HEADER_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOODWE_FILL = PatternFill("solid", fgColor="EAF3F8")
FRONIUS_FILL = PatternFill("solid", fgColor="FFF4D6")
FIMER_FILL = PatternFill("solid", fgColor="EDEAF8")
SOLIS_FILL = PatternFill("solid", fgColor="EAF6EA")
SOLAX_FILL = PatternFill("solid", fgColor="EAF2FF")
SUMMARY_FILL = PatternFill("solid", fgColor="EEF5F1")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)


def capacity_from_name(name):
    match = re.search(r"(\d+(?:\.\d+)?)\s*k\s*w", name or "", re.IGNORECASE)
    return float(match.group(1)) if match else None


def write_csv(path, headers, rows):
    lines = [",".join(headers)]
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header)
            if value is None:
                value = ""
            value = str(value).replace('"', '""')
            values.append(f'"{value}"' if "," in value or "\n" in value else value)
        lines.append(",".join(values))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def read_optional_json(path):
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else None


def add_solis_rows(rows, solis):
    if not solis:
        return
    for system in solis.get("systems", []):
        rows.append(
            {
                "Brand": "Solis",
                "System Name": system.get("name") or system.get("station_name"),
                "Status": system.get("status"),
                "Capacity (kW)": system.get("capacity_kw") or system.get("capacity"),
                "Current Power (kW)": system.get("current_power_kw"),
                "Today Generation (kWh)": system.get("today_generation_kwh"),
                "Weekly Generation (kWh)": system.get("weekly_generation_kwh"),
                "Total Generation (kWh)": system.get("total_generation_kwh"),
                "Source ID": system.get("system_id") or system.get("station_id"),
            }
        )


def add_solax_rows(rows, solax):
    if not solax:
        return
    for system in solax.get("systems", []):
        rows.append(
            {
                "Brand": "SolaX",
                "System Name": system.get("name") or system.get("station_name"),
                "Status": system.get("status"),
                "Capacity (kW)": system.get("capacity_kw") or system.get("capacity"),
                "Current Power (kW)": system.get("current_power_kw"),
                "Today Generation (kWh)": system.get("today_generation_kwh"),
                "Weekly Generation (kWh)": system.get("weekly_generation_kwh"),
                "Total Generation (kWh)": system.get("total_generation_kwh"),
                "Source ID": system.get("system_id") or system.get("station_id"),
            }
        )


def brand_fill(brand):
    if brand == "GoodWe":
        return GOODWE_FILL
    if brand == "Fronius":
        return FRONIUS_FILL
    if brand == "FIMER":
        return FIMER_FILL
    if brand == "Solis":
        return SOLIS_FILL
    if brand == "SolaX":
        return SOLAX_FILL
    return PatternFill("solid", fgColor="F6F8FA")


def brand_summary_rows(rows):
    brands = []
    for row in rows:
        if row["Brand"] not in brands:
            brands.append(row["Brand"])

    summary = []
    for brand in brands:
        brand_rows = [row for row in rows if row["Brand"] == brand]
        today_total = sum(row["Today Generation (kWh)"] or 0 for row in brand_rows)
        weekly_total = sum(row["Weekly Generation (kWh)"] or 0 for row in brand_rows)
        summary.append(
            (
                f"{brand} Today kWh",
                today_total,
                f"{brand} Systems",
                len(brand_rows),
            )
        )
        summary.append(
            (
                f"{brand} Weekly kWh",
                weekly_total,
                "Report Type" if brand == brands[-1] else "",
                "Static Numbers-compatible" if brand == brands[-1] else "",
            )
        )
    summary.append(("Total Systems", len(rows), "Brands", len(brands)))
    return summary


def main():
    goodwe = json.loads(GOODWE_FILE.read_text(encoding="utf-8"))
    goodwe_weekly = json.loads(GOODWE_WEEKLY_FILE.read_text(encoding="utf-8"))
    fronius_weekly = json.loads(FRONIUS_WEEKLY_FILE.read_text(encoding="utf-8"))
    fronius_current = json.loads(FRONIUS_CURRENT_FILE.read_text(encoding="utf-8"))
    fimer = read_optional_json(FIMER_FILE)
    solis = read_optional_json(SOLIS_FILE)
    solax = read_optional_json(SOLAX_FILE)

    goodwe_weekly_by_id = {
        station["station_id"]: station for station in goodwe_weekly["stations"]
    }
    fronius_current_by_id = {
        system["system_id"]: system for system in fronius_current["systems"]
    }

    rows = []
    for station in goodwe["stations"]:
        weekly = goodwe_weekly_by_id.get(station.get("powerstation_id"), {})
        rows.append(
            {
                "Brand": "GoodWe",
                "System Name": station.get("stationname"),
                "Status": station.get("status"),
                "Capacity (kW)": station.get("capacity"),
                "Current Power (kW)": station.get("pac_kw"),
                "Today Generation (kWh)": station.get("eday"),
                "Weekly Generation (kWh)": weekly.get("weekly_generation_kwh"),
                "Total Generation (kWh)": station.get("etotal"),
                "Source ID": station.get("powerstation_id"),
            }
        )

    for system in fronius_weekly["systems"]:
        current = fronius_current_by_id.get(system.get("system_id"), {})
        rows.append(
            {
                "Brand": "Fronius",
                "System Name": system.get("name"),
                "Status": system.get("status"),
                "Capacity (kW)": capacity_from_name(system.get("name")),
                "Current Power (kW)": None,
                "Today Generation (kWh)": current.get("today_generation_kwh"),
                "Weekly Generation (kWh)": system.get("weekly_generation_kwh"),
                "Total Generation (kWh)": current.get("total_generation_kwh"),
                "Source ID": system.get("system_id"),
            }
        )

    if fimer:
        for item in fimer.get("plantEnergy", []):
            plant = item["plant"]

            def energy_value(key):
                value = item.get("values", {}).get(key, {})
                body = value.get("body", [])
                if value.get("status") == 200 and body:
                    return body[0].get("value")
                return None

            rows.append(
                {
                    "Brand": "FIMER",
                    "System Name": plant.get("name"),
                    "Status": plant.get("state"),
                    "Capacity (kW)": plant.get("configuration", {}).get(
                        "panelsNominalPower"
                    ),
                    "Current Power (kW)": None,
                    "Today Generation (kWh)": energy_value("today"),
                    "Weekly Generation (kWh)": energy_value("week"),
                    "Total Generation (kWh)": energy_value("total"),
                    "Source ID": plant.get("entityID"),
                }
            )

    add_solis_rows(rows, solis)
    add_solax_rows(rows, solax)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Generation Report"
    sheet.sheet_view.showGridLines = False

    brands = []
    for row in rows:
        if row["Brand"] not in brands:
            brands.append(row["Brand"])
    sheet["A1"] = " + ".join(brands) + " Generation Report"
    sheet["A1"].font = Font(bold=True, size=18, color="17324D")
    sheet["A2"] = (
        f"GoodWe weekly: {goodwe_weekly['start_date']} to {goodwe_weekly['end_date']} | "
        f"Fronius weekly: {fronius_weekly['start_date']} to {fronius_weekly['end_date']}"
    )
    sheet["A2"].font = Font(color="576575")

    summary_rows = brand_summary_rows(rows)
    for index, values in enumerate(summary_rows, start=4):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=index, column=column, value=value)
            cell.fill = SUMMARY_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if column in (1, 3):
                cell.font = Font(bold=True, color="17324D")
            if column in (2,):
                cell.number_format = "#,##0.00"

    headers = list(rows[0].keys())
    table_start = 4 + len(summary_rows) + 2
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=table_start, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row_data in enumerate(rows, start=table_start + 1):
        fill = brand_fill(row_data["Brand"])
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row_data.get(header))
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(header == "System Name"))
            if column_index in (4, 5, 6, 7, 8):
                cell.number_format = "#,##0.00"

    last_row = table_start + len(rows)

    chart_data_start = last_row + 3
    sheet.cell(row=chart_data_start, column=1, value="Chart Data").font = Font(
        bold=True, color="17324D"
    )
    chart_headers = ["System Name", "Weekly Generation (kWh)"]
    for column_index, header in enumerate(chart_headers, start=1):
        cell = sheet.cell(row=chart_data_start + 1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    sorted_rows = sorted(rows, key=lambda row: row["Weekly Generation (kWh)"] or 0, reverse=True)
    for row_index, row_data in enumerate(sorted_rows, start=chart_data_start + 2):
        sheet.cell(row=row_index, column=1, value=f"{row_data['Brand']} - {row_data['System Name']}")
        value_cell = sheet.cell(row=row_index, column=2, value=row_data["Weekly Generation (kWh)"] or 0)
        value_cell.number_format = "#,##0.00"
        for column_index in (1, 2):
            sheet.cell(row=row_index, column=column_index).border = THIN_BORDER

    chart = BarChart()
    chart.title = "Weekly Generation by System"
    chart.y_axis.title = "kWh"
    chart.height = 10
    chart.width = 20
    values = Reference(
        sheet,
        min_col=2,
        min_row=chart_data_start + 1,
        max_row=chart_data_start + 1 + len(sorted_rows),
    )
    categories = Reference(
        sheet,
        min_col=1,
        min_row=chart_data_start + 2,
        max_row=chart_data_start + 1 + len(sorted_rows),
    )
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    sheet.add_chart(chart, "K4")

    widths = {
        "A": 18,
        "B": 32,
        "C": 14,
        "D": 15,
        "E": 18,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 40,
        "K": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A11"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    write_csv(CSV_FILE, headers, rows)
    print(f"Saved Numbers-compatible workbook to {OUTPUT_FILE}")
    print(f"Saved CSV fallback to {CSV_FILE}")


if __name__ == "__main__":
    main()
