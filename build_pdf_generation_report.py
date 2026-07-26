import datetime as dt
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


SOURCE_FILE = Path("outputs/numbers_compatible/numbers_compatible_generation_report.xlsx")
OUTPUT_FILE = Path("outputs/numbers_compatible/numbers_compatible_generation_report.pdf")


def short(value, max_len=34):
    text = "" if value is None else str(value)
    return text if len(text) <= max_len else text[: max_len - 3] + "..."


def number(value):
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return str(value)


def load_report_data():
    workbook = load_workbook(SOURCE_FILE, data_only=True)
    sheet = workbook["Generation Report"]

    title = sheet["A1"].value or "Generation Report"
    subtitle = sheet["A2"].value or ""

    summary = []
    row = 4
    while sheet.cell(row, 1).value and sheet.cell(row, 1).value != "Brand":
        left_label = sheet.cell(row, 1).value
        left_value = sheet.cell(row, 2).value
        right_label = sheet.cell(row, 3).value
        right_value = sheet.cell(row, 4).value
        if left_label:
            summary.append((left_label, left_value, right_label, right_value))
        row += 1

    header_row = row
    while sheet.cell(header_row, 1).value != "Brand":
        header_row += 1
        if header_row > sheet.max_row:
            raise RuntimeError("Could not find report table header row")
    headers = [sheet.cell(header_row, column).value for column in range(1, 10)]
    rows = []
    row_index = header_row + 1
    while sheet.cell(row_index, 1).value:
        if sheet.cell(row_index, 1).value == "Chart Data":
            break
        rows.append([sheet.cell(row_index, column).value for column in range(1, 10)])
        row_index += 1

    return title, subtitle, summary, headers, rows


def build_chart_rows(headers, rows):
    weekly_index = headers.index("Weekly Generation (kWh)")
    totals = {}
    for row in rows:
        brand = row[0] or "Unknown"
        totals[brand] = totals.get(brand, 0) + (row[weekly_index] or 0)
    max_value = max(totals.values()) if totals else 0
    chart_rows = [["Brand", "Weekly kWh", "Scale"]]
    for brand, value in sorted(totals.items(), key=lambda item: item[1], reverse=True):
        width = int((value / max_value) * 34) if max_value else 0
        chart_rows.append([brand, number(value), "#" * max(width, 1)])
    return chart_rows


def add_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#576575"))
    canvas.drawString(15 * mm, 9 * mm, f"Generated {dt.datetime.now():%Y-%m-%d %H:%M}")
    canvas.drawRightString(282 * mm, 9 * mm, f"Page {doc.page}")
    canvas.restoreState()


def main():
    title, subtitle, summary, headers, rows = load_report_data()
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#17324D"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#576575"),
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        alignment=TA_LEFT,
    )
    center_style = ParagraphStyle(
        "CenterCell",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    doc = SimpleDocTemplate(
        str(OUTPUT_FILE),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=14 * mm,
        title=title,
    )

    story = [
        Paragraph(title, title_style),
        Paragraph(short(subtitle, 180), subtitle_style),
    ]

    summary_data = []
    for left_label, left_value, right_label, right_value in summary:
        summary_data.append(
            [
                Paragraph(str(left_label), cell_style),
                Paragraph(number(left_value), center_style),
                Paragraph(str(right_label or ""), cell_style),
                Paragraph(number(right_value) if isinstance(right_value, (int, float)) else str(right_value or ""), center_style),
            ]
        )
    summary_table = Table(summary_data, colWidths=[44 * mm, 30 * mm, 42 * mm, 45 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF5F1")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7DEE8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 8)])

    chart_data = build_chart_rows(headers, rows)
    chart_table = Table(chart_data, colWidths=[32 * mm, 30 * mm, 88 * mm])
    chart_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F6F8FA")),
                ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor("#2E6F9E")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7DEE8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE8")),
                ("FONTNAME", (2, 1), (2, -1), "Courier"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.extend([Paragraph("Weekly Generation by Brand", subtitle_style), chart_table, PageBreak()])

    visible_indexes = [
        index
        for index, header in enumerate(headers)
        if header and header != "Source ID"
    ]
    visible_headers = [headers[index] for index in visible_indexes]
    table_data = [[Paragraph(str(header), center_style) for header in visible_headers]]
    for row in rows:
        values = []
        for index in visible_indexes:
            header = headers[index]
            value = row[index]
            if isinstance(value, (int, float)) and "Status" not in header:
                values.append(number(value))
            else:
                values.append(short(value, 30 if header == "System Name" else 14))
        table_data.append([Paragraph(value, cell_style) for value in values])

    data_table = LongTable(
        table_data,
        repeatRows=1,
        colWidths=[18 * mm, 54 * mm, 22 * mm, 22 * mm, 27 * mm, 30 * mm, 32 * mm, 34 * mm][
            : len(visible_headers)
        ],
    )
    data_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FAFBFC")),
                ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7DEE8")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D7DEE8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FA")]),
            ]
        )
    )
    story.extend([Paragraph("System Details", subtitle_style), data_table])

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    print(f"Saved PDF report to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
