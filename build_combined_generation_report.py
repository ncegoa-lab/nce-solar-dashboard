import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


GOODWE_FILE = Path("sems_station_data.json")
GOODWE_WEEKLY_FILE = Path("sems_weekly_generation.json")
FRONIUS_FILE = Path("fronius_weekly_generation.json")
FRONIUS_CURRENT_FILE = Path("fronius_current_generation.json")
OUTPUT_DIR = Path("outputs/combined_generation")
OUTPUT_FILE = OUTPUT_DIR / "combined_goodwe_fronius_generation_report.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="17324D")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GOODWE_FILL = PatternFill("solid", fgColor="EAF3F8")
FRONIUS_FILL = PatternFill("solid", fgColor="FFF4D6")
SUMMARY_FILL = PatternFill("solid", fgColor="EEF5F1")
THIN_BORDER = Border(
    left=Side(style="thin", color="D7DEE8"),
    right=Side(style="thin", color="D7DEE8"),
    top=Side(style="thin", color="D7DEE8"),
    bottom=Side(style="thin", color="D7DEE8"),
)


def capacity_from_name(name):
    match = re.search(r"(\d+(?:\.\d+)?)\s*k\s*w", name, re.IGNORECASE)
    return float(match.group(1)) if match else None


def main():
    goodwe = json.loads(GOODWE_FILE.read_text(encoding="utf-8"))
    goodwe_weekly = json.loads(GOODWE_WEEKLY_FILE.read_text(encoding="utf-8"))
    fronius = json.loads(FRONIUS_FILE.read_text(encoding="utf-8"))
    fronius_current = json.loads(FRONIUS_CURRENT_FILE.read_text(encoding="utf-8"))
    fronius_current_by_id = {
        system["system_id"]: system for system in fronius_current["systems"]
    }
    goodwe_weekly_by_id = {
        station["station_id"]: station for station in goodwe_weekly["stations"]
    }

    rows = []
    for station in goodwe["stations"]:
        weekly = goodwe_weekly_by_id.get(station.get("powerstation_id"), {})
        rows.append(
            {
                "Brand": "GoodWe",
                "System Name": station.get("stationname"),
                "Status": station.get("status"),
                "Capacity (kW)": station.get("capacity"),
                "Current Power (kW)": station.get("pac_kw"),
                "Today Generation (kWh)": station.get("eday"),
                "Weekly Generation (kWh)": weekly.get("weekly_generation_kwh"),
                "Total Generation (kWh)": station.get("etotal"),
                "Source ID": station.get("powerstation_id"),
            }
        )

    for system in fronius["systems"]:
        current = fronius_current_by_id.get(system.get("system_id"), {})
        rows.append(
            {
                "Brand": "Fronius",
                "System Name": system.get("name"),
                "Status": system.get("status"),
                "Capacity (kW)": capacity_from_name(system.get("name", "")),
                "Current Power (kW)": None,
                "Today Generation (kWh)": current.get("today_generation_kwh"),
                "Weekly Generation (kWh)": system.get("weekly_generation_kwh"),
                "Total Generation (kWh)": current.get("total_generation_kwh"),
                "Source ID": system.get("system_id"),
            }
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Combined Report"
    sheet.sheet_view.showGridLines = False

    sheet["A1"] = "GoodWe + Fronius Generation Report"
    sheet["A1"].font = Font(bold=True, size=18, color="17324D")
    sheet["A2"] = (
        f"GoodWe snapshot: {goodwe['generated_at']} | "
        f"Fronius weekly period: {fronius['start_date']} to {fronius['end_date']} | "
        f"Fronius current date: {fronius_current['date']}"
    )
    sheet["A2"].font = Font(color="576575")

    sheet["A4"] = "GoodWe Today kWh"
    sheet["B4"] = "=SUMIF(A10:A200,\"GoodWe\",F10:F200)"
    sheet["A5"] = "GoodWe Weekly kWh"
    sheet["B5"] = "=SUMIF(A10:A200,\"GoodWe\",G10:G200)"
    sheet["A6"] = "Fronius Weekly kWh"
    sheet["B6"] = "=SUMIF(A10:A200,\"Fronius\",G10:G200)"
    sheet["A7"] = "Fronius Total kWh"
    sheet["B7"] = "=SUMIF(A10:A200,\"Fronius\",H10:H200)"
    sheet["D4"] = "GoodWe Systems"
    sheet["E4"] = "=COUNTIF(A10:A200,\"GoodWe\")"
    sheet["D5"] = "Fronius Systems"
    sheet["E5"] = "=COUNTIF(A10:A200,\"Fronius\")"
    for row in sheet.iter_rows(min_row=4, max_row=7, min_col=1, max_col=5):
        for cell in row:
            cell.fill = SUMMARY_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
    for cell in ("A4", "A5", "A6", "A7", "D4", "D5"):
        sheet[cell].font = Font(bold=True, color="17324D")
    sheet["B4"].number_format = "#,##0.00"
    sheet["B5"].number_format = "#,##0.00"
    sheet["B6"].number_format = "#,##0.00"
    sheet["B7"].number_format = "#,##0.00"

    headers = list(rows[0].keys())
    start_row = 9
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=column_index, value=header)

    for cell in sheet[start_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_index, row_data in enumerate(rows, start=start_row + 1):
        for column_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=column_index, value=row_data[header])

    last_row = start_row + len(rows)
    for row in sheet.iter_rows(min_row=start_row + 1, max_row=last_row, min_col=1, max_col=len(headers)):
        brand = row[0].value
        fill = GOODWE_FILL if brand == "GoodWe" else FRONIUS_FILL
        for cell in row:
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        for cell in row[3:8]:
            cell.number_format = "#,##0.00"

    widths = {
        "A": 12,
        "B": 30,
        "C": 14,
        "D": 15,
        "E": 18,
        "F": 22,
        "G": 22,
        "H": 22,
        "I": 40,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    goodwe_chart = BarChart()
    goodwe_chart.title = "GoodWe Weekly Generation (kWh)"
    goodwe_chart.y_axis.title = "kWh"
    goodwe_values = Reference(sheet, min_col=7, min_row=start_row, max_row=start_row + len(goodwe["stations"]))
    goodwe_categories = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=start_row + len(goodwe["stations"]))
    goodwe_chart.add_data(goodwe_values, titles_from_data=True)
    goodwe_chart.set_categories(goodwe_categories)
    goodwe_chart.height = 8
    goodwe_chart.width = 16
    sheet.add_chart(goodwe_chart, "K4")

    fronius_start = start_row + len(goodwe["stations"]) + 1
    fronius_end = last_row
    fronius_chart = BarChart()
    fronius_chart.title = "Fronius Weekly Generation (kWh)"
    fronius_chart.y_axis.title = "kWh"
    fronius_values = Reference(sheet, min_col=7, min_row=start_row, max_row=fronius_end)
    fronius_categories = Reference(sheet, min_col=2, min_row=start_row + 1, max_row=fronius_end)
    fronius_chart.add_data(fronius_values, titles_from_data=True)
    fronius_chart.set_categories(fronius_categories)
    fronius_chart.height = 8
    fronius_chart.width = 16
    sheet.add_chart(fronius_chart, "K20")

    sheet.freeze_panes = "A10"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_FILE)
    print(f"Saved combined report to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
