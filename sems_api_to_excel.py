import datetime as dt
import json
import os
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook


LOGIN_URL = "https://www.semsportal.com/api/v2/Common/CrossLogin"
OUTPUT_FILE = Path("sems_station_snapshot.xlsx")
HEADERS = [
    "Timestamp",
    "Station Name",
    "Station ID",
    "Status",
    "Capacity (kW)",
    "Current Power (kW)",
    "Today Generation (kWh)",
    "Month Generation (kWh)",
    "Total Generation (kWh)",
    "Location",
    "Organization",
]


def require_env(name: str) -> str:
    value = os.getenv(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def token_header(token_data: Optional[dict[str, object]] = None) -> str:
    header = {
        "version": "v2.1.0",
        "client": "ios",
        "language": "en",
    }
    if token_data:
        for key in ("uid", "timestamp", "token"):
            if key in token_data:
                header[key] = token_data[key]
    return json.dumps(header, separators=(",", ":"))


def login(session: requests.Session) -> tuple[str, dict[str, object]]:
    response = session.post(
        LOGIN_URL,
        json={
            "account": require_env("SEMS_USERNAME"),
            "pwd": require_env("SEMS_PASSWORD"),
        },
        timeout=30,
    )
    response.raise_for_status()
    body = response.json()
    if body.get("hasError"):
        raise RuntimeError(f"SEMS login failed: {body.get('msg')}")
    return body.get("api", "https://www.semsportal.com/api/"), body["data"]


def get_stations(session: requests.Session, api_base: str) -> list[dict[str, object]]:
    stations = []
    page_index = 1
    page_size = 100
    total_records = None

    while total_records is None or len(stations) < total_records:
        response = session.post(
            f"{api_base}v2/PowerStationMonitor/QueryPowerStationMonitor",
            json={
                "powerStationId": "",
                "key": "",
                "pageIndex": page_index,
                "pageSize": page_size,
            },
            timeout=30,
        )
        response.raise_for_status()
        body = response.json()
        if body.get("hasError"):
            raise RuntimeError(f"Station fetch failed: {body.get('msg')}")

        data = body.get("data", {})
        page_stations = data.get("list", [])
        total_records = data.get("record", len(page_stations))
        stations.extend(page_stations)

        if not page_stations:
            break
        page_index += 1

    return stations


def append_rows(rows: list[dict[str, object]]) -> None:
    if OUTPUT_FILE.exists():
        workbook = load_workbook(OUTPUT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SEMS Snapshot"
        sheet.append(HEADERS)

    for row in rows:
        sheet.append([row.get(header) for header in HEADERS])

    workbook.save(OUTPUT_FILE)


def main() -> None:
    session = requests.Session()
    session.headers.update(
        {
            "Content-Type": "application/json",
            "Token": token_header(),
            "User-Agent": "Mozilla/5.0",
        }
    )

    api_base, token_data = login(session)
    session.headers.update({"Token": token_header(token_data)})

    timestamp = dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")
    stations = get_stations(session, api_base)
    rows = [
        {
            "Timestamp": timestamp,
            "Station Name": station.get("stationname"),
            "Station ID": station.get("powerstation_id"),
            "Status": station.get("status"),
            "Capacity (kW)": station.get("capacity"),
            "Current Power (kW)": station.get("pac_kw"),
            "Today Generation (kWh)": station.get("eday"),
            "Month Generation (kWh)": station.get("emonth"),
            "Total Generation (kWh)": station.get("etotal"),
            "Location": station.get("location"),
            "Organization": station.get("org_name"),
        }
        for station in stations
    ]

    append_rows(rows)
    print(f"Added {len(rows)} station rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
